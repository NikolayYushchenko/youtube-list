import openpyxl

def write_dict_keys_to_excel(data_dict, filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        if(workbook):
            sheet = workbook.active
            header = [cell.value for cell in sheet[1]]
            keys_data_dict = list(data_dict.keys())

            if keys_data_dict != header:
                print(f"Заголовки таблиці були змінені: {header}!!! Потрібні заголовки: {keys_data_dict}")
                for col_idx, col_name in enumerate(keys_data_dict, start=1):
                    sheet.cell(row=1, column=col_idx, value=col_name)
                workbook.save(filename)
                print("Заголовки таблиці переписані!")

        return workbook

    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        column_names = list(data_dict.keys())
        sheet.append(column_names)
        workbook.save(filename)
        print("Файл створено і заголовки додані в таблицю!")
        return workbook

def append_dict_to_excel(data_dict, filename):

    workbook = write_dict_keys_to_excel(data_dict, filename)
    sheet = workbook.active

    last_row = sheet.max_row + 1

    values = [data_dict.get(col_name, '') for col_name in data_dict.keys()]
    for col_idx, value in enumerate(values, start=1):
        sheet.cell(row=last_row, column=col_idx, value=value)

    workbook.save(filename)
    print("Дані додані.")

def read_dicts_from_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Назва стовпчиків
    column_names = [cell.value for cell in sheet[1]]

    # Читаємо значення і формуємо список словників
    data_dicts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_dict = {}
        for col_index, value in enumerate(row):
            data_dict[column_names[col_index]] = value
        data_dicts.append(data_dict)

    return data_dicts

def find_and_replace_row_by_value(filename, search_value, new_data_dict):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Зчитуємо перший рядок
    column_names = [cell.value for cell in sheet[1]]

    found = False  # Прапор для перевірки считування ключового слова

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if search_value in row:
            found = True
            for col_index, col_name in enumerate(column_names):
                new_value = new_data_dict.get(col_name, '')
                sheet.cell(row=row_number, column=col_index + 1, value=new_value)

    if found:
        workbook.save(filename)
        print("Рядки перезаписані успішно.")
    else:
        print("Значення ключового слова не знайдено в рядках.")

def get_column_data_by_index(filename, column_index):

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    column_number = column_index  # Номер колонки
    column_data = []

    # Проходим по рядкам і получаємо дані
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Начинаем с второй строки, предполагая, что первая строка - заголовки
        if row[column_number - 1]:  # Проверяем, что значение в ячейке не пустое
            column_data.append(row[column_number - 1])

    workbook.close()
    return column_data


def get_column_values_by_name(file_path, column_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    column_values = []

    # Получаем индекс столбца по имени
    column_index = None
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=1):
        for cell in col:
            if cell.value == column_name:
                column_index = cell.column
                break

    if column_index:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
            column_values.append(row[0].value)

    return column_values

def write_value_to_cell_by_names(file_path, row_name, column_name, value):
    # Открываем файл Excel
    workbook = openpyxl.load_workbook(file_path)

    # Выбираем активный лист (или можно указать имя листа)
    sheet = workbook.active

    # Получаем номер столбца по имени колонки
    column_number = None
    for col_idx, col in enumerate(sheet.iter_cols(min_row=1, max_row=1), start=1):
        if col[0].value == column_name:
            column_number = col_idx
            break

    if column_number is None:
        print(f"Колонка {column_name} не найдена.")
        return

    # Находим номер строки по имени ряда
    row_number = None
    for row_idx, row in enumerate(sheet.iter_rows(min_col=1, max_col=1), start=1):
        if row[0].value == row_name:
            row_number = row_idx
            break

    if row_number is None:
        print(f"Ряд {row_name} не найден.")
        return

    # Записываем значение в соответствующую ячейку
    cell = sheet.cell(row=row_number, column=column_number)
    cell.value = value

    # Сохраняем изменения
    workbook.save(file_path)
    workbook.close()

#print(get_column_data_by_index("uma_data.xlsx", 1))
print(get_column_values_by_name("uma_data.xlsx", 'ID'))


"""
file_path = 'uma_data.xlsx'
row_name = '909'
column_name = 'Status'
value = 'New Value'

write_value_to_cell_by_names(file_path, row_name, column_name, value)

new_data = {
    'Name': 'John1',
    'Age': 31,
    'Country': 'USA1'
}


search_value = 'John1'

new_row_data = {
    'Name': 'John1',
    'Age': 32393495959549594,
    'Country': 'UK1'
}


newaa_row_data = {
    'Name': 'John',
    'Age': 32,
    'Country': 'UK'
}


# Дозапис словника
append_dict_to_excel(new_data, 'data.xlsx')
append_dict_to_excel(newaa_row_data, 'data.xlsx')

# Читання з файлу
all_data = read_dicts_from_excel('data.xlsx')
#print(all_data)
for data_dict in all_data:
    print(data_dict)

# Заміна по ключу старих значень на нові в рядках співпадань
find_and_replace_row_by_value('data.xlsx', search_value, new_row_data)

# Читання з файлу
all_data = read_dicts_from_excel('data.xlsx')
#print(all_data)
for data_dict in all_data:
    print(data_dict)
"""
